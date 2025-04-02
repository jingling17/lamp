import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

st.set_page_config(page_title="落地灯台灯数据分析工具", layout="wide")

def load_data(file):
    """加载Excel数据"""
    try:
        df = pd.read_excel(file)
        required_columns = ['商品名称', '商品链接', '零售额', '零售量', '成交均价', '品牌']
        if not all(col in df.columns for col in required_columns):
            st.error("Excel文件格式不正确，请确保包含所有必需列：商品名称、商品链接、零售额、零售量、成交均价、品牌")
            return None
        return df
    except Exception as e:
        st.error(f"加载文件时出错：{str(e)}")
        return None

def combine_platform_data(uploaded_files, platform_names, period_names):
    """合并多个平台和多个时间段的数据"""
    all_data = []
    
    for file, platform, period in zip(uploaded_files, platform_names, period_names):
        df = load_data(file)
        if df is not None:
            df['平台'] = platform
            df['时间段'] = period
            all_data.append(df)
    
    if len(all_data) > 0:
        combined_df = pd.concat(all_data, ignore_index=True)
        return combined_df
    else:
        st.error("所有上传的文件均无法解析，请检查文件格式")
        return None

def calculate_period_stats(df, period, platform_col='平台'):
    """计算时间段统计信息"""
    # 按时间段和平台分组计算统计数据
    stats = df.groupby([period, platform_col]).agg({
        '零售额': 'sum',
        '零售量': 'sum',
        '成交均价': 'mean'
    }).round(2).reset_index()
    
    # 添加平台累计数据
    total_stats = df.groupby([period]).agg({
        '零售额': 'sum',
        '零售量': 'sum',
        '成交均价': 'mean'
    }).round(2).reset_index()
    total_stats[platform_col] = '所有平台'
    
    # 合并平台数据和累计数据
    all_stats = pd.concat([stats, total_stats], ignore_index=True)
    
    # 计算同比变化
    # 为每个平台单独计算同比变化
    result_dfs = []
    
    for platform in all_stats[platform_col].unique():
        platform_data = all_stats[all_stats[platform_col] == platform].copy().sort_values(by=period)
        platform_data['零售额同比变化'] = platform_data['零售额'].pct_change() * 100
        platform_data['零售量同比变化'] = platform_data['零售量'].pct_change() * 100
        platform_data['成交均价同比变化'] = platform_data['成交均价'].pct_change() * 100
        result_dfs.append(platform_data)
    
    final_stats = pd.concat(result_dfs, ignore_index=True)
    
    return final_stats

def calculate_brand_share(df, period, platform_col='平台'):
    """计算品牌占比"""
    # 按时间段、平台和品牌分组计算统计数据
    brand_stats = df.groupby([period, platform_col, '品牌']).agg({
        '零售额': 'sum',
        '零售量': 'sum'
    }).reset_index()
    
    # 计算每个时间段和平台的总额和总量
    period_platform_totals = df.groupby([period, platform_col]).agg({
        '零售额': 'sum',
        '零售量': 'sum'
    })
    
    # 计算占比
    brand_stats['零售额占比'] = brand_stats.apply(
        lambda x: (x['零售额'] / period_platform_totals.loc[(x[period], x[platform_col]), '零售额']) * 100, 
        axis=1
    )
    brand_stats['零售量占比'] = brand_stats.apply(
        lambda x: (x['零售量'] / period_platform_totals.loc[(x[period], x[platform_col]), '零售量']) * 100, 
        axis=1
    )
    
    # 添加所有平台合计数据
    all_platform_brand_stats = df.groupby([period, '品牌']).agg({
        '零售额': 'sum',
        '零售量': 'sum'
    }).reset_index()
    
    all_platform_totals = df.groupby([period]).agg({
        '零售额': 'sum',
        '零售量': 'sum'
    })
    
    all_platform_brand_stats[platform_col] = '所有平台'
    all_platform_brand_stats['零售额占比'] = all_platform_brand_stats.apply(
        lambda x: (x['零售额'] / all_platform_totals.loc[x[period], '零售额']) * 100, 
        axis=1
    )
    all_platform_brand_stats['零售量占比'] = all_platform_brand_stats.apply(
        lambda x: (x['零售量'] / all_platform_totals.loc[x[period], '零售量']) * 100, 
        axis=1
    )
    
    # 合并所有数据
    all_brand_stats = pd.concat([brand_stats, all_platform_brand_stats], ignore_index=True)
    
    return all_brand_stats

def analyze_price_segments(df, period, price_ranges, platform_col='平台'):
    """分析价位段"""
    # 创建价位段标签
    ranges = price_ranges
    df = df.copy()
    labels = [f"{ranges[i]}-{ranges[i+1]}" for i in range(len(ranges)-1)]
    df['价位段'] = pd.cut(df['成交均价'], bins=price_ranges, labels=labels)
    
    # 按时间段、平台和价位段分组计算统计数据
    segment_stats = df.groupby([period, platform_col, '价位段']).agg({
        '零售额': 'sum',
        '零售量': 'sum'
    }).reset_index()
    
    # 计算每个时间段和平台的总额和总量
    period_platform_totals = df.groupby([period, platform_col]).agg({
        '零售额': 'sum',
        '零售量': 'sum'
    })
    
    # 计算占比
    segment_stats['零售额占比'] = segment_stats.apply(
        lambda x: (x['零售额'] / period_platform_totals.loc[(x[period], x[platform_col]), '零售额']) * 100 if x['价位段'] is not None else 0, 
        axis=1
    )
    segment_stats['零售量占比'] = segment_stats.apply(
        lambda x: (x['零售量'] / period_platform_totals.loc[(x[period], x[platform_col]), '零售量']) * 100 if x['价位段'] is not None else 0, 
        axis=1
    )
    
    # 添加所有平台合计数据
    all_platform_segment_stats = df.groupby([period, '价位段']).agg({
        '零售额': 'sum',
        '零售量': 'sum'
    }).reset_index()
    
    all_platform_totals = df.groupby([period]).agg({
        '零售额': 'sum',
        '零售量': 'sum'
    })
    
    all_platform_segment_stats[platform_col] = '所有平台'
    all_platform_segment_stats['零售额占比'] = all_platform_segment_stats.apply(
        lambda x: (x['零售额'] / all_platform_totals.loc[x[period], '零售额']) * 100 if x['价位段'] is not None else 0,
        axis=1
    )
    all_platform_segment_stats['零售量占比'] = all_platform_segment_stats.apply(
        lambda x: (x['零售量'] / all_platform_totals.loc[x[period], '零售量']) * 100 if x['价位段'] is not None else 0,
        axis=1
    )
    
    # 合并所有数据
    all_segment_stats = pd.concat([segment_stats, all_platform_segment_stats], ignore_index=True)
    
    # 计算同比变化
    result_dfs = []
    
    for platform in all_segment_stats[platform_col].unique():
        platform_data = all_segment_stats[(all_segment_stats[platform_col] == platform)].copy()
        
        for price_segment in platform_data['价位段'].unique():
            if price_segment is not None:
                segment_data = platform_data[platform_data['价位段'] == price_segment].sort_values(by=period)
                segment_data['零售额占比变化'] = segment_data['零售额占比'].pct_change() * 100
                segment_data['零售量占比变化'] = segment_data['零售量占比'].pct_change() * 100
                result_dfs.append(segment_data)
    
    if result_dfs:
        final_segment_stats = pd.concat(result_dfs, ignore_index=True)
        return final_segment_stats
    else:
        return all_segment_stats

def get_top_brands_by_segment(df, period, price_ranges, platform_col='平台', n=5):
    """获取每个价位段的TOP品牌"""
    # 创建价位段标签
    ranges = price_ranges
    df = df.copy()
    labels = [f"{ranges[i]}-{ranges[i+1]}" for i in range(len(ranges)-1)]
    df['价位段'] = pd.cut(df['成交均价'], bins=price_ranges, labels=labels)
    
    # 按时间段、平台、价位段和品牌分组计算统计数据
    top_brands = df.groupby([period, platform_col, '价位段', '品牌']).agg({
        '零售额': 'sum',
        '零售量': 'sum'
    }).reset_index()
    
    # 计算每个时间段、平台和价位段的总量
    segment_totals = df.groupby([period, platform_col, '价位段']).agg({
        '零售额': 'sum',
        '零售量': 'sum'
    })
    
    # 计算品牌占比
    def calculate_share(row):
        try:
            if row['价位段'] is not None:
                total = segment_totals.loc[(row[period], row[platform_col], row['价位段'])]
                return (row['零售额'] / total['零售额']) * 100
            return 0
        except:
            return 0
    
    def calculate_volume_share(row):
        try:
            if row['价位段'] is not None:
                total = segment_totals.loc[(row[period], row[platform_col], row['价位段'])]
                return (row['零售量'] / total['零售量']) * 100
            return 0
        except:
            return 0
    
    top_brands['零售额占比'] = top_brands.apply(calculate_share, axis=1)
    top_brands['零售量占比'] = top_brands.apply(calculate_volume_share, axis=1)
    
    # 获取各平台TOP N品牌
    result_dfs = []
    
    for p in top_brands[period].unique():
        for platform in top_brands[platform_col].unique():
            for segment in top_brands['价位段'].unique():
                if segment is not None:
                    segment_data = top_brands[
                        (top_brands[period] == p) & 
                        (top_brands[platform_col] == platform) & 
                        (top_brands['价位段'] == segment)
                    ].sort_values(by='零售额', ascending=False).head(n)
                    
                    result_dfs.append(segment_data)
    
    # 添加所有平台合计数据
    # 按时间段、价位段和品牌分组计算汇总数据
    all_platform_top_brands = df.groupby([period, '价位段', '品牌']).agg({
        '零售额': 'sum',
        '零售量': 'sum'
    }).reset_index()
    
    # 计算每个时间段和价位段的总量
    all_platform_segment_totals = df.groupby([period, '价位段']).agg({
        '零售额': 'sum',
        '零售量': 'sum'
    })
    
    # 计算汇总的品牌占比
    def calculate_all_platform_share(row):
        try:
            if row['价位段'] is not None:
                total = all_platform_segment_totals.loc[(row[period], row['价位段'])]
                return (row['零售额'] / total['零售额']) * 100
            return 0
        except:
            return 0
    
    def calculate_all_platform_volume_share(row):
        try:
            if row['价位段'] is not None:
                total = all_platform_segment_totals.loc[(row[period], row['价位段'])]
                return (row['零售量'] / total['零售量']) * 100
            return 0
        except:
            return 0
    
    all_platform_top_brands['零售额占比'] = all_platform_top_brands.apply(calculate_all_platform_share, axis=1)
    all_platform_top_brands['零售量占比'] = all_platform_top_brands.apply(calculate_all_platform_volume_share, axis=1)
    all_platform_top_brands[platform_col] = '所有平台'
    
    # 获取所有平台TOP N品牌
    all_platform_result_dfs = []
    
    for p in all_platform_top_brands[period].unique():
        for segment in all_platform_top_brands['价位段'].unique():
            if segment is not None:
                segment_data = all_platform_top_brands[
                    (all_platform_top_brands[period] == p) & 
                    (all_platform_top_brands['价位段'] == segment)
                ].sort_values(by='零售额', ascending=False).head(n)
                
                all_platform_result_dfs.append(segment_data)
    
    # 合并所有结果
    if result_dfs and all_platform_result_dfs:
        top_brands_results = pd.concat(result_dfs + all_platform_result_dfs, ignore_index=True)
        return top_brands_results
    elif result_dfs:
        return pd.concat(result_dfs, ignore_index=True)
    elif all_platform_result_dfs:
        return pd.concat(all_platform_result_dfs, ignore_index=True)
    else:
        return pd.DataFrame()

def get_top_products_by_segment(df, period, price_ranges, platform_col='平台', n=5):
    """获取每个价位段的TOP产品"""
    # 创建价位段标签
    ranges = price_ranges
    df = df.copy()
    labels = [f"{ranges[i]}-{ranges[i+1]}" for i in range(len(ranges)-1)]
    df['价位段'] = pd.cut(df['成交均价'], bins=price_ranges, labels=labels)
    
    # 按时间段、平台、价位段和产品名称分组计算统计数据
    top_products = df.groupby([period, platform_col, '价位段', '商品名称', '商品链接']).agg({
        '零售额': 'sum',
        '零售量': 'sum',
        '成交均价': 'mean'
    }).reset_index()
    
    # 获取各平台TOP N产品
    result_dfs = []
    
    for p in top_products[period].unique():
        for platform in top_products[platform_col].unique():
            for segment in top_products['价位段'].unique():
                if segment is not None:
                    segment_data = top_products[
                        (top_products[period] == p) & 
                        (top_products[platform_col] == platform) & 
                        (top_products['价位段'] == segment)
                    ].sort_values(by='零售额', ascending=False).head(n)
                    
                    result_dfs.append(segment_data)
    
    # 添加所有平台合计数据
    # 按时间段、价位段和产品名称分组计算汇总数据
    all_platform_top_products = df.groupby([period, '价位段', '商品名称', '商品链接']).agg({
        '零售额': 'sum',
        '零售量': 'sum',
        '成交均价': 'mean'
    }).reset_index()
    
    all_platform_top_products[platform_col] = '所有平台'
    
    # 获取所有平台TOP N产品
    all_platform_result_dfs = []
    
    for p in all_platform_top_products[period].unique():
        for segment in all_platform_top_products['价位段'].unique():
            if segment is not None:
                segment_data = all_platform_top_products[
                    (all_platform_top_products[period] == p) & 
                    (all_platform_top_products['价位段'] == segment)
                ].sort_values(by='零售额', ascending=False).head(n)
                
                all_platform_result_dfs.append(segment_data)
    
    # 合并所有结果
    if result_dfs and all_platform_result_dfs:
        top_products_results = pd.concat(result_dfs + all_platform_result_dfs, ignore_index=True)
        return top_products_results
    elif result_dfs:
        return pd.concat(result_dfs, ignore_index=True)
    elif all_platform_result_dfs:
        return pd.concat(all_platform_result_dfs, ignore_index=True)
    else:
        return pd.DataFrame()

def plot_period_comparison(df, period, metric, platform_col='平台'):
    """绘制时间段对比图表"""
    fig = px.line(df, x=period, y=metric, color=platform_col,
                  title=f'{metric}时间段对比',
                  labels={period: '时间段', metric: metric})
    return fig

def plot_brand_comparison(df, period, platform_col='平台', n=10):
    """绘制品牌对比图表"""
    # 获取每个时间段TOP N品牌
    top_brands = df.groupby([period, platform_col, '品牌'])['零售额'].sum().reset_index()
    top_brands = top_brands.sort_values([period, platform_col, '零售额'], ascending=[True, True, False])
    top_brands = top_brands.groupby([period, platform_col]).head(n)
    
    fig = px.bar(top_brands, x='品牌', y='零售额', color=platform_col,
                 title=f'TOP {n}品牌对比',
                 labels={'品牌': '品牌名称', '零售额': '零售额'})
    return fig

def plot_price_segment_comparison(df, period, platform_col='平台'):
    """绘制价位段对比图表"""
    fig = px.bar(df, x='价位段', y='零售额占比', color=platform_col,
                 title='各价位段零售额占比对比',
                 labels={'价位段': '价位段', '零售额占比': '零售额占比(%)'})
    return fig

def format_excel_output(writer, period_stats, top_10_brands, segment_stats, top_brands, top_products):
    """美化Excel输出格式，让不同时间段对比更直观"""
    # 创建样式
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
    
    # 基础样式
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # 安全处理分类数据和NaN值
    # 1. 对于包含分类数据的DataFrame，先将其转换为普通DataFrame
    if hasattr(period_stats, 'dtypes'):
        for col, dtype in period_stats.dtypes.items():
            if str(dtype) == 'category':
                period_stats[col] = period_stats[col].astype(str)
    
    if hasattr(top_10_brands, 'dtypes'):
        for col, dtype in top_10_brands.dtypes.items():
            if str(dtype) == 'category':
                top_10_brands[col] = top_10_brands[col].astype(str)
    
    if hasattr(segment_stats, 'dtypes'):
        for col, dtype in segment_stats.dtypes.items():
            if str(dtype) == 'category':
                segment_stats[col] = segment_stats[col].astype(str)
    
    if hasattr(top_brands, 'dtypes'):
        for col, dtype in top_brands.dtypes.items():
            if str(dtype) == 'category':
                top_brands[col] = top_brands[col].astype(str)
    
    if hasattr(top_products, 'dtypes'):
        for col, dtype in top_products.dtypes.items():
            if str(dtype) == 'category':
                top_products[col] = top_products[col].astype(str)
    
    # 2. 然后再填充NaN值
    period_stats = period_stats.fillna(0)
    top_10_brands = top_10_brands.fillna(0)
    segment_stats = segment_stats.fillna(0)
    top_brands = top_brands.fillna(0)
    top_products = top_products.fillna(0)
    
    # 确保至少写入一个工作表，防止"At least one sheet must be visible"错误
    # 写入基础数据表，这些表必定会存在
    period_stats.to_excel(writer, sheet_name='时间段统计', index=False)
    top_10_brands.to_excel(writer, sheet_name='品牌占比分析', index=False)
    segment_stats.to_excel(writer, sheet_name='价位段统计', index=False)
    
    # 写入可能为空的表
    if not top_brands.empty:
        top_brands.to_excel(writer, sheet_name='价位段TOP品牌', index=False)
    else:
        # 如果表为空，创建一个简单的表防止错误
        pd.DataFrame({'提示': ['没有找到符合条件的价位段TOP品牌数据']}).to_excel(
            writer, sheet_name='价位段TOP品牌', index=False)
    
    if not top_products.empty:
        top_products.to_excel(writer, sheet_name='价位段TOP产品', index=False)
    else:
        # 如果表为空，创建一个简单的表防止错误
        pd.DataFrame({'提示': ['没有找到符合条件的价位段TOP产品数据']}).to_excel(
            writer, sheet_name='价位段TOP产品', index=False)
    
    # 创建各种对比专用表
    create_period_comparison_sheet(writer, period_stats)
    create_brand_comparison_sheet(writer, top_10_brands)
    create_segment_comparison_sheet(writer, segment_stats)
    
    # 只有当数据不为空时才创建这些表
    if not top_brands.empty and len(top_brands['时间段'].unique()) > 1:
        create_top_brands_comparison_sheet(writer, top_brands)
    
    if not top_products.empty and len(top_products['时间段'].unique()) > 1:
        create_top_products_comparison_sheet(writer, top_products)
    
    # 应用样式到所有工作表
    for sheet_name in writer.sheets:
        ws = writer.sheets[sheet_name]
        
        # 设置表头样式
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 调整列宽
        for i, col in enumerate(ws.columns, 1):
            # 调整列宽时考虑内容长度
            max_length = 0
            for cell in col:
                if cell.value:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
            adjusted_width = min(max(15, max_length + 2), 40)  # 最小15，最大40
            ws.column_dimensions[get_column_letter(i)].width = adjusted_width
        
        # 应用条件格式到变化率列
        if sheet_name == '时间段统计':
            for col in ['零售额同比变化', '零售量同比变化', '成交均价同比变化']:
                if col in period_stats.columns:
                    try:
                        col_idx = list(period_stats.columns).index(col) + 1
                        ws.conditional_formatting.add(
                            f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{len(period_stats)+1}",
                            ColorScaleRule(
                                start_type='min', start_color='E67C73',  # 红色（负值）
                                mid_type='num', mid_value=0, mid_color='FFFFFF',  # 白色（零值）
                                end_type='max', end_color='57BB8A'  # 绿色（正值）
                            )
                        )
                    except Exception as e:
                        pass  # 忽略条件格式错误，不影响主要功能
        
        # 对所有对比分析表应用条件格式
        if '对比分析' in sheet_name:
            try:
                # 对比表使用更明显的条件格式
                for col_idx in range(2, ws.max_column + 1):
                    col_letter = get_column_letter(col_idx)
                    # 只对变化率列应用条件格式
                    if ws[f'{col_letter}1'].value and ('(%)' in str(ws[f'{col_letter}1'].value) or '变化' in str(ws[f'{col_letter}1'].value)):
                        ws.conditional_formatting.add(
                            f"{col_letter}2:{col_letter}{ws.max_row}",
                            ColorScaleRule(
                                start_type='min', start_color='E67C73',
                                mid_type='num', mid_value=0, mid_color='FFFFFF',
                                end_type='max', end_color='57BB8A'
                            )
                        )
            except Exception as e:
                pass  # 忽略条件格式错误，不影响主要功能
            
            # 添加交替行颜色，增强可读性
            try:
                for row in range(2, ws.max_row + 1, 2):
                    for col in range(1, ws.max_column + 1):
                        cell = ws[f'{get_column_letter(col)}{row}']
                        cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            except Exception as e:
                pass  # 忽略样式错误，不影响主要功能

def create_period_comparison_sheet(writer, period_stats):
    """创建时间段对比专用表"""
    # 提取所有时间段
    periods = sorted(period_stats['时间段'].unique())  # 确保时间段是排序的
    platforms = period_stats['平台'].unique()
    
    # 只有当有多个时间段时才创建对比表
    if len(periods) <= 1:
        return
        
    # 创建对比数据
    comparison_data = []
    
    for platform in platforms:
        platform_data = period_stats[period_stats['平台'] == platform]
        
        if len(platform_data) > 1:
            # 对每个平台计算不同时间段之间的变化
            for metric in ['零售额', '零售量', '成交均价']:
                row = {'指标': f'{platform} - {metric}'}
                
                # 添加每个时间段的原始值
                for period in periods:
                    period_value = platform_data[platform_data['时间段'] == period][metric].values
                    if len(period_value) > 0:
                        row[period] = period_value[0]
                    else:
                        row[period] = None
                
                # 计算环比变化率
                for i in range(1, len(periods)):
                    prev_period = periods[i-1]
                    curr_period = periods[i]
                    
                    if prev_period in row and curr_period in row and row[prev_period] is not None and row[curr_period] is not None and row[prev_period] != 0:
                        row[f'{curr_period} vs {prev_period} (%)'] = ((row[curr_period] / row[prev_period]) - 1) * 100
                    else:
                        row[f'{curr_period} vs {prev_period} (%)'] = None
                
                comparison_data.append(row)
    
    # 创建DataFrame并写入Excel
    if comparison_data:
        comparison_df = pd.DataFrame(comparison_data)
        
        # 重新排列列顺序，保证先显示所有时间段的数据，再显示环比变化
        columns = ['指标']
        for period in periods:
            columns.append(period)
        for i in range(1, len(periods)):
            columns.append(f'{periods[i]} vs {periods[i-1]} (%)')
        
        # 确保列在DataFrame中存在再进行选择
        final_columns = ['指标']
        for col in columns[1:]:
            if col in comparison_df.columns:
                final_columns.append(col)
        
        # 重新排序列
        comparison_df = comparison_df[final_columns]
        
        comparison_df.to_excel(writer, sheet_name='时间段对比分析', index=False)
        
        # 格式化百分比列
        ws = writer.sheets['时间段对比分析']
        for i, col_name in enumerate(comparison_df.columns):
            if '(%)' in col_name:
                col_letter = get_column_letter(i + 1)
                for row in range(2, len(comparison_df) + 2):
                    cell = ws[f'{col_letter}{row}']
                    if cell.value is not None:
                        cell.number_format = '0.00%'
            elif col_name in periods:  # 为数值列添加千分位格式
                col_letter = get_column_letter(i + 1)
                for row in range(2, len(comparison_df) + 2):
                    cell = ws[f'{col_letter}{row}']
                    if cell.value is not None and '均价' in ws['A'+str(row)].value:
                        cell.number_format = '#,##0.00'  # 均价保留两位小数
                    elif cell.value is not None:
                        cell.number_format = '#,##0'  # 其他数值采用千分位格式

def create_brand_comparison_sheet(writer, top_brands):
    """创建品牌占比对比专用表"""
    # 如果没有品牌数据或时间段不足，则不创建
    if top_brands.empty or len(top_brands['时间段'].unique()) <= 1:
        return
    
    # 按平台和品牌分组，展示各时间段的占比变化
    comparison_data = []
    
    for platform in top_brands['平台'].unique():
        platform_data = top_brands[top_brands['平台'] == platform]
        
        # 获取该平台下所有时间段的TOP品牌
        top_brands_in_platform = platform_data['品牌'].unique()
        periods = platform_data['时间段'].unique()
        
        for brand in top_brands_in_platform:
            brand_data = platform_data[platform_data['品牌'] == brand]
            
            if len(brand_data) > 1:  # 只分析在多个时间段出现的品牌
                row = {'平台': platform, '品牌': brand}
                
                # 添加每个时间段的零售额占比
                for period in periods:
                    period_value = brand_data[brand_data['时间段'] == period]
                    if not period_value.empty:
                        row[f'{period} 零售额'] = period_value['零售额'].values[0]
                        row[f'{period} 零售额占比(%)'] = period_value['零售额占比'].values[0]
                    else:
                        row[f'{period} 零售额'] = None
                        row[f'{period} 零售额占比(%)'] = None
                
                # 计算环比变化
                prev_share = None
                prev_period = None
                
                for period in sorted(periods):
                    share_key = f'{period} 零售额占比(%)'
                    if share_key in row and row[share_key] is not None:
                        if prev_share is not None and prev_period is not None:
                            row[f'{period} vs {prev_period} 占比变化(%)'] = row[share_key] - prev_share
                        prev_share = row[share_key]
                        prev_period = period
                
                comparison_data.append(row)
    
    # 创建DataFrame并写入Excel
    if comparison_data:
        comparison_df = pd.DataFrame(comparison_data)
        # 按平台和零售额占比降序排序
        comparison_df.to_excel(writer, sheet_name='品牌占比对比分析', index=False)

def create_segment_comparison_sheet(writer, segment_stats):
    """创建价位段对比专用表"""
    # 如果没有价位段数据或时间段不足，则不创建
    if segment_stats.empty or len(segment_stats['时间段'].unique()) <= 1:
        return
    
    # 按平台和价位段分组，展示各时间段的占比变化
    comparison_data = []
    
    for platform in segment_stats['平台'].unique():
        platform_data = segment_stats[segment_stats['平台'] == platform]
        
        # 获取该平台下所有价位段
        segments = platform_data['价位段'].unique()
        periods = platform_data['时间段'].unique()
        
        for segment in segments:
            if segment is not None:
                segment_data = platform_data[platform_data['价位段'] == segment]
                
                if len(segment_data) > 1:  # 只分析在多个时间段出现的价位段
                    row = {'平台': platform, '价位段': segment}
                    
                    # 添加每个时间段的零售额占比
                    for period in periods:
                        period_value = segment_data[segment_data['时间段'] == period]
                        if not period_value.empty:
                            row[f'{period} 零售额'] = period_value['零售额'].values[0]
                            row[f'{period} 零售额占比(%)'] = period_value['零售额占比'].values[0]
                        else:
                            row[f'{period} 零售额'] = None
                            row[f'{period} 零售额占比(%)'] = None
                    
                    # 计算环比变化
                    prev_share = None
                    prev_period = None
                    
                    for period in sorted(periods):
                        share_key = f'{period} 零售额占比(%)'
                        if share_key in row and row[share_key] is not None:
                            if prev_share is not None and prev_period is not None:
                                row[f'{period} vs {prev_period} 占比变化(%)'] = row[share_key] - prev_share
                            prev_share = row[share_key]
                            prev_period = period
                    
                    comparison_data.append(row)
    
    # 创建DataFrame并写入Excel
    if comparison_data:
        comparison_df = pd.DataFrame(comparison_data)
        comparison_df.to_excel(writer, sheet_name='价位段占比对比分析', index=False)

def create_top_brands_comparison_sheet(writer, top_brands_segment):
    """创建价位段TOP品牌对比专用表"""
    # 如果没有价位段TOP品牌数据或时间段不足，则不创建
    if top_brands_segment.empty or len(top_brands_segment['时间段'].unique()) <= 1:
        return
    
    # 按平台、价位段和品牌分组，展示各时间段的占比变化
    comparison_data = []
    
    for platform in top_brands_segment['平台'].unique():
        platform_data = top_brands_segment[top_brands_segment['平台'] == platform]
        
        for segment in platform_data['价位段'].unique():
            if segment is not None:
                segment_data = platform_data[platform_data['价位段'] == segment]
                
                # 获取该价位段下所有品牌
                brands = segment_data['品牌'].unique()
                periods = segment_data['时间段'].unique()
                
                for brand in brands:
                    brand_data = segment_data[segment_data['品牌'] == brand]
                    
                    if len(brand_data) > 1:  # 只分析在多个时间段出现的品牌
                        row = {'平台': platform, '价位段': segment, '品牌': brand}
                        
                        # 添加每个时间段的零售额占比
                        for period in sorted(periods):
                            period_value = brand_data[brand_data['时间段'] == period]
                            if not period_value.empty:
                                row[f'{period} 零售额'] = period_value['零售额'].values[0]
                                row[f'{period} 零售额占比(%)'] = period_value['零售额占比'].values[0]
                            else:
                                row[f'{period} 零售额'] = None
                                row[f'{period} 零售额占比(%)'] = None
                        
                        # 计算环比变化
                        prev_share = None
                        prev_period = None
                        
                        for period in sorted(periods):
                            share_key = f'{period} 零售额占比(%)'
                            if share_key in row and row[share_key] is not None:
                                if prev_share is not None and prev_period is not None:
                                    row[f'{period} vs {prev_period} 占比变化(%)'] = row[share_key] - prev_share
                                prev_share = row[share_key]
                                prev_period = period
                        
                        comparison_data.append(row)
    
    # 创建DataFrame并写入Excel
    if comparison_data:
        comparison_df = pd.DataFrame(comparison_data)
        
        # 按价位段排序品牌对比数据
        try:
            # 获取最后一个时间段
            periods = sorted(top_brands_segment['时间段'].unique())
            if periods:
                last_period = periods[-1]
                last_period_column = f'{last_period} 零售额占比(%)'
                
                # 如果存在这个列，按它排序
                if last_period_column in comparison_df.columns:
                    comparison_df = comparison_df.sort_values(
                        by=['平台', '价位段', last_period_column], 
                        ascending=[True, True, False]
                    )
        except:
            # 如果排序出错，保持原有顺序
            pass
            
        comparison_df.to_excel(writer, sheet_name='价位段品牌对比分析', index=False)

def create_top_products_comparison_sheet(writer, top_products):
    """创建价位段TOP产品时间段对比专用表"""
    # 如果没有价位段产品数据或时间段不足，则不创建
    if top_products.empty or len(top_products['时间段'].unique()) <= 1:
        return
    
    # 按平台、价位段和商品分组，展示各时间段的销量和均价变化
    comparison_data = []
    
    # 跟踪每个产品在所有时间段的出现情况
    product_period_count = {}
    
    # 首先统计每个产品在不同时间段的出现次数
    for platform in top_products['平台'].unique():
        platform_data = top_products[top_products['平台'] == platform]
        
        for segment in platform_data['价位段'].unique():
            if segment is not None:
                segment_data = platform_data[platform_data['价位段'] == segment]
                
                for _, row in segment_data.iterrows():
                    product_key = (platform, segment, row['商品名称'])
                    if product_key not in product_period_count:
                        product_period_count[product_key] = set()
                    product_period_count[product_key].add(row['时间段'])
    
    # 创建热销产品的时间段对比数据
    for platform in top_products['平台'].unique():
        platform_data = top_products[top_products['平台'] == platform]
        
        for segment in platform_data['价位段'].unique():
            if segment is not None:
                segment_data = platform_data[platform_data['价位段'] == segment]
                
                # 获取该价位段下所有产品
                product_names = segment_data['商品名称'].unique()
                periods = segment_data['时间段'].unique()
                
                # 按照时间段数量排序，优先展示在多个时间段都出现的产品
                sorted_products = sorted(
                    [(name, len(product_period_count.get((platform, segment, name), set()))) 
                     for name in product_names],
                    key=lambda x: x[1], 
                    reverse=True
                )
                
                for product_name, count in sorted_products:
                    if count > 1:  # 只对比在多个时间段出现的产品
                        product_data = segment_data[segment_data['商品名称'] == product_name]
                        
                        row = {
                            '平台': platform, 
                            '价位段': segment, 
                            '商品名称': product_name, 
                            '商品链接': product_data['商品链接'].iloc[0]
                        }
                        
                        # 添加每个时间段的零售额、零售量和成交均价
                        for period in periods:
                            period_value = product_data[product_data['时间段'] == period]
                            if not period_value.empty:
                                row[f'{period} 零售额'] = period_value['零售额'].values[0]
                                row[f'{period} 零售量'] = period_value['零售量'].values[0]
                                row[f'{period} 成交均价'] = period_value['成交均价'].values[0]
                            else:
                                row[f'{period} 零售额'] = None
                                row[f'{period} 零售量'] = None
                                row[f'{period} 成交均价'] = None
                        
                        # 计算环比变化
                        sorted_periods = sorted(periods)
                        
                        # 零售额变化
                        prev_value = None
                        prev_period = None
                        
                        for period in sorted_periods:
                            value_key = f'{period} 零售额'
                            if value_key in row and row[value_key] is not None:
                                if prev_value is not None and prev_period is not None and prev_value != 0:
                                    row[f'{period} vs {prev_period} 零售额变化(%)'] = ((row[value_key] / prev_value) - 1) * 100
                                prev_value = row[value_key]
                                prev_period = period
                        
                        # 零售量变化
                        prev_value = None
                        prev_period = None
                        
                        for period in sorted_periods:
                            value_key = f'{period} 零售量'
                            if value_key in row and row[value_key] is not None:
                                if prev_value is not None and prev_period is not None and prev_value != 0:
                                    row[f'{period} vs {prev_period} 零售量变化(%)'] = ((row[value_key] / prev_value) - 1) * 100
                                prev_value = row[value_key]
                                prev_period = period
                        
                        # 成交均价变化
                        prev_value = None
                        prev_period = None
                        
                        for period in sorted_periods:
                            value_key = f'{period} 成交均价'
                            if value_key in row and row[value_key] is not None:
                                if prev_value is not None and prev_period is not None and prev_value != 0:
                                    row[f'{period} vs {prev_period} 均价变化(%)'] = ((row[value_key] / prev_value) - 1) * 100
                                prev_value = row[value_key]
                                prev_period = period
                        
                        comparison_data.append(row)
    
    # 创建DataFrame并写入Excel
    if comparison_data:
        comparison_df = pd.DataFrame(comparison_data)
        
        # 按平台、价位段排序，并按零售额降序排序
        last_period = sorted(top_products['时间段'].unique())[-1] if len(top_products['时间段'].unique()) > 0 else None
        if last_period and f'{last_period} 零售额' in comparison_df.columns:
            comparison_df = comparison_df.sort_values(
                by=['平台', '价位段', f'{last_period} 零售额'], 
                ascending=[True, True, False]
            )
        
        comparison_df.to_excel(writer, sheet_name='价位段产品对比分析', index=False)

def main():
    st.title("落地灯台灯数据分析工具")
    
    # 平台和时间段信息
    st.subheader("数据上传配置")
    
    col1, col2 = st.columns(2)
    
    with col1:
        num_platforms = st.number_input("上传平台数量", min_value=1, max_value=10, value=1)
    
    with col2:
        num_periods = st.number_input("上传时间段数量", min_value=1, max_value=10, value=1)
    
    uploaded_files = []
    platform_names = []
    period_names = []
    
    # 时间段命名
    st.write("时间段命名")
    period_labels = []
    for i in range(num_periods):
        period_label = st.text_input(f"时间段 {i+1} 名称", f"时间段 {i+1}")
        period_labels.append(period_label)
    
    # 平台配置和数据上传
    for i in range(num_platforms):
        st.write(f"平台 {i+1} 配置")
        col1, col2 = st.columns(2)
        
        with col1:
            platform_name = st.text_input(f"平台 {i+1} 名称", f"平台 {i+1}")
        
        # 为每个平台上传多个时间段的数据
        for j in range(num_periods):
            uploaded_file = st.file_uploader(
                f"上传 {platform_name} - {period_labels[j]} Excel数据文件", 
                type=['xlsx', 'xls'], 
                key=f"platform_{i}_period_{j}"
            )
            uploaded_files.append(uploaded_file)
            platform_names.append(platform_name)
            period_names.append(period_labels[j])
    
    # 过滤掉未上传的文件
    valid_files = [(f, p, t) for f, p, t in zip(uploaded_files, platform_names, period_names) if f is not None]
    uploaded_valid_files = [f[0] for f in valid_files]
    platform_valid_names = [f[1] for f in valid_files]
    period_valid_names = [f[2] for f in valid_files]
    
    if len(valid_files) > 0:
        # 合并所有平台和时间段数据
        df = combine_platform_data(uploaded_valid_files, platform_valid_names, period_valid_names)
        
        if df is not None:
            # 显示已上传的数据时间段
            st.write("已上传的时间段：", ", ".join(sorted(df['时间段'].unique())))
            
            # 添加自定义价位段配置
            st.subheader("价位段配置")
            min_price = float(df['成交均价'].min())
            max_price = float(df['成交均价'].max())
            
            with st.expander("自定义价位段"):
                use_custom_segments = st.checkbox("使用自定义价位段", value=False)
                
                if use_custom_segments:
                    # 用户自定义价位段界限
                    st.write("输入自定义价位段界限（小值到大值，用逗号分隔）：")
                    custom_ranges_str = st.text_input("例如：0,100,300,500,1000", f"{min_price},{(min_price+max_price)/2},{max_price}")
                    
                    try:
                        price_ranges = [float(x) for x in custom_ranges_str.split(',')]
                        price_ranges.sort()  # 确保是升序
                        
                        # 检查是否至少有两个界限
                        if len(price_ranges) < 2:
                            st.error("至少需要输入两个价位值才能形成一个价位段。")
                            price_ranges = [min_price, max_price]
                    except:
                        st.error("输入格式有误，请使用逗号分隔的数字。")
                        price_ranges = [min_price, max_price]
                        
                else:
                    # 自动生成价位段
                    num_segments = st.number_input("设置价位段数量", min_value=2, max_value=10, value=5)
                    price_ranges = [min_price]
                    
                    for i in range(num_segments - 1):
                        price = st.number_input(f"价位段 {i+1} 上限", 
                                              min_value=float(min_price), 
                                              max_value=float(max_price),
                                              value=float(min_price + (max_price - min_price) * (i + 1) / num_segments))
                        price_ranges.append(price)
                    price_ranges.append(max_price)
                
                # 显示最终价位段
                segments_display = [f"{price_ranges[i]}-{price_ranges[i+1]}" for i in range(len(price_ranges)-1)]
                st.write("最终价位段设置：", ", ".join(segments_display))
            
            # 分析按钮
            if st.button("开始分析"):
                st.session_state.analysis_done = True
                
                # 1. 时间段统计
                st.subheader("1. 时间段统计")
                period_stats = calculate_period_stats(df, '时间段')
                
                # 添加时间段对比图表
                st.write("时间段对比图表")
                metrics = ['零售额', '零售量', '成交均价']
                for metric in metrics:
                    fig = plot_period_comparison(period_stats, '时间段', metric)
                    st.plotly_chart(fig, use_container_width=True)
                
                st.dataframe(period_stats)
                
                # 2. 品牌占比分析
                st.subheader("2. TOP10品牌占比分析")
                brand_stats = calculate_brand_share(df, '时间段')
                
                # 添加品牌对比图表
                st.write("品牌对比图表")
                fig = plot_brand_comparison(brand_stats, '时间段')
                st.plotly_chart(fig, use_container_width=True)
                
                # 获取每个时间段和平台的TOP10品牌
                top_10_brands_list = []
                
                for p in brand_stats['时间段'].unique():
                    for platform in brand_stats['平台'].unique():
                        platform_period_data = brand_stats[
                            (brand_stats['时间段'] == p) & 
                            (brand_stats['平台'] == platform)
                        ].sort_values(by='零售额', ascending=False).head(10)
                        
                        top_10_brands_list.append(platform_period_data)
                
                top_10_brands = pd.concat(top_10_brands_list, ignore_index=True)
                st.dataframe(top_10_brands)
                
                # 3. 价位段分析
                st.subheader("3. 价位段分析")
                
                # 3a. 价位段统计
                st.write("3a. 价位段统计")
                segment_stats = analyze_price_segments(df, '时间段', price_ranges)
                
                # 添加价位段对比图表
                st.write("价位段对比图表")
                fig = plot_price_segment_comparison(segment_stats, '时间段')
                st.plotly_chart(fig, use_container_width=True)
                
                st.dataframe(segment_stats)
                
                # 3b. 价位段TOP品牌
                st.write("3b. 价位段TOP5品牌")
                top_brands = get_top_brands_by_segment(df, '时间段', price_ranges)
                st.dataframe(top_brands)
                
                # 3c. 价位段TOP产品
                st.write("3c. 价位段TOP5产品")
                top_products = get_top_products_by_segment(df, '时间段', price_ranges)
                st.dataframe(top_products)
                
                # 导出分析结果
                st.subheader("导出分析结果")
                
                with pd.ExcelWriter('分析结果.xlsx', engine='openpyxl') as writer:
                    # 使用美化功能
                    format_excel_output(writer, period_stats, top_10_brands, segment_stats, top_brands, top_products)
                
                with open('分析结果.xlsx', 'rb') as f:
                    st.download_button(
                        label="下载分析结果",
                        data=f,
                        file_name="分析结果.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

if __name__ == "__main__":
    main() 